from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import matplotlib.image as mpimg

output_file = "../data/housing location data.xlsx"
sheet_name = "Housing location data"


file_input = "../data/Check_in_Cleaned.xlsx"
main_camp_img = mpimg.imread("../data/images/main campus map.png")
main_camp_img = "main"
bobcat_village_img = mpimg.imread("../data/images/bobcat village.png")
bobcat_village_img = "bobcat_village"
housing_list = {
    "Alamito Hall": {
        "x": 1295.56,
        "y": 140.81,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Bexar Hall": {
        "x": 713.23,
        "y": 634.31,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Blanco Hall": {
        "x": 343.10,
        "y": 565.22,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Bobcat Village Apartments": {
        "x": 217.40,
        "y": 212.48,
        "dist": 0.7,
        "size": 1,
        "image": bobcat_village_img
    },
    "Brogdon Hall": {
        "x": 1495.42,
        "y": 412.24,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Butler Hall": {
        "x": 1574.39,
        "y": 547.95,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Richard A. Castro Hall": {
        "x": 737.90,
        "y": 493.66,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Chautauqua Hall": {
        "x": 851.41,
        "y": 185.22,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Cibolo Hall": {
        "x": 1265.95,
        "y": 209.90,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "College Inn": {
        "x": 876.08,
        "y": 306.13,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Elena Zamora O'Shea Hall": {
        "x": 165.44,
        "y": 244.45,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Falls Hall": {
        "x": 397.39,
        "y": 570.16,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "First Five Freedom Hall": {
        "x": 195.05,
        "y": 313.54,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Gaillardia Hall": {
        "x": 821.80,
        "y": 244.45,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Jackson Hall": {
        "x": 1058.68,
        "y": 229.64,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Lantana Hall": {
        "x": 1522.57,
        "y": 560.29,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Laurel Hall": {
        "x": 1431.27,
        "y": 493.66,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Mesquite Hall": {
        "x": 1426.33,
        "y": 424.57,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Retama Hall": {
        "x": 1379.45,
        "y": 488.73,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "San Jacinto Hall": {
        "x": 1120.36,
        "y": 599.77,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "San Marcos Hall": {
        "x": 658.94,
        "y": 557.82,
        "dist": 0.7,
        "image": main_camp_img
    },
    "Sayers Hall": {
        "x": 439.33,
        "y": 636.78,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Sterry Hall": {
        "x": 1485.56,
        "y": 639.25,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Tower Hall": {
        "x": 1194.39,
        "y": 560.29,
        "dist": 0.7,
        "size": 1,
        "image": main_camp_img
    },
    "Derrick Hall": {
        "x": 1281.19,
        "y": 296.62,
        "dist": 0.7,
        "size": 200,
        "image": main_camp_img
    }
}

df = pd.read_excel(file_input)


num_students = df.groupby("Housing Type").size()

num_unique_students = df.groupby("Housing Type")["NetID"].nunique()

total_on_campus_students = num_students.sum()
normalized_value = num_students / total_on_campus_students

# Put into a summary df
summary = pd.DataFrame({
    "NumStudents": num_students,
    "NumUniqueStudents": num_unique_students,
    "NormalizedStudents": normalized_value
})

rows = []
for housing_name, info in housing_list.items():
    rows.append({
        "Housing": housing_name,
        "X": info["x"],
        "Y": info["y"],
        "Dist": info["dist"],
        "ImageType": info["image"]
    })

df = pd.DataFrame(rows)

final_df = df.merge(summary, how="left", left_on="Housing", right_index=True)

final_df.to_excel(output_file, sheet_name=sheet_name, index=False)

wb = load_workbook(output_file)
ws = wb[sheet_name]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheet_name)
wb.save(output_file)