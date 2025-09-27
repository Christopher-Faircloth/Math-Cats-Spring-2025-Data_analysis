#this file looks through weeks of mathcats to see which students came in consistantly throughout the semester
#consistantly is defined by once a week for at least 5 week since first coming and discovering mathcats
#this can be sued to see grade comparison of sutdents who were consistant vs weren't
#note: this definition can be changed to something like every other week, or every week -1 (leaneance)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = "../data/Check_In_Cleaned.xlsx"
output_file = "../results/consistant_students.xlsx"
sheetName = "consistant students"

original = pd.read_excel(input_file)
original['Check In Date'] = pd.to_datetime(original['Check In Date'])

original['Week'] = original['Check In Date'].dt.isocalendar().week

# spring break shouldnt count
spring_break_start = pd.to_datetime("2025-03-09")
spring_break_end   = pd.to_datetime("2025-03-16")

spring_break_weeks = set(original.loc[
    (original['Check In Date'] >= spring_break_start) &
    (original['Check In Date'] <= spring_break_end),
    'Week'
].unique())

# many students dont need to come the last week because they already finished their finals
last_week = original['Week'].max()

valid_weeks = set(original['Week'].unique()) - spring_break_weeks - {last_week}

def consistent_after_first(weeks):
    weeks = sorted(set(map(int, weeks)))
    first = weeks[0]  # their first attendance
    required = {w for w in valid_weeks if w >= first}
    return required.issubset(weeks)

student_weeks = original.groupby('NetID')['Week'].unique()
consistent_students = student_weeks[student_weeks.apply(consistent_after_first)]
#at least came for 5 weeks 
consistent_students = consistent_students[consistent_students.apply(len) >= 5]


df_consistent = consistent_students.apply(
    lambda x: ", ".join(map(str, sorted(map(int, x))))
).reset_index()

df_consistent.columns = ["NetID", "Weeks"]

df_consistent.to_excel(output_file, sheet_name = sheetName, index=False)

# format the excel file
wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)



