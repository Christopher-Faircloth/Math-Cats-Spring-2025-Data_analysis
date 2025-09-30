# This files adds new columns and catagorization so that further regression analysis can be looked at
# new coulmns; weekday, hour time, major category, and course category

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_grades = "../data/Students Grade Report.csv"
file_check_in = "../data/Check_In_Cleaned.xlsx"
file_output = "../data/Check_In_Cleaned.xlsx"
sheetName = "check in cleaned"
    
df = pd.read_excel(file_check_in)
df_grades = pd.read_csv(file_grades)

df["NetID"] = df["NetID"].astype(str)
df_grades["NetID"] = df_grades["NetID"].astype(str)

grades = df_grades[["NetID", "Enrollment Course Number", "Midterm Grade", "Final Grade"]]

merged = pd.merge(
    df, 
    grades,
    left_on = ["NetID", "Course Number"],
    right_on = ["NetID", "Enrollment Course Number"],
    how = "inner")

# Save to a new Excel file
merged.to_excel(file_output, sheet_name = sheetName, index=False)
# format the excel file
wb = load_workbook(file_output)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(file_output)



