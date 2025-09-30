import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_output = "../data/Check_In_Cleaned.xlsx"
sheetName = "check in cleaned"

df = pd.read_excel(file_output)
#anonymizes student names
uniqueStudents = df["NetID"].unique()
counter = 1000
for name in uniqueStudents:
    df.loc[df["NetID"] == name, "NetID"] = counter
    counter += 1

# Save to a new Excel file
df.to_excel(file_output, sheet_name = sheetName, index=False)
# format the excel file
wb = load_workbook(file_output)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(file_output)



