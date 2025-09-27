# This files adds new columns and catagorization so that further regression analysis can be looked at
# new coulmns; weekday, hour time, major category, and course category

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

file_input = "../data/Full Spring 2025.xlsx"
file_output = "../data/Check_In_Cleaned.xlsx"
sheetName = "check in cleaned"

# catagories computer science, calculus, stats, algebra, precal, proof-based, Math (other), science, other
# catagory filters
def categorize_course(row):
    course_number = str(row["Course Number"])
    course_name = str(row["Course Name"])
    course_name = course_name.lower()
    if course_number.startswith("CS-"):
        return "Computer Science"
    # cal 1,2,3 business and econ 2, cal life science 1,2, integral calculus
    elif course_number in ["MATH-2471", "MATH-2472", "MATH-2393", "MATH-2321", "MATH-2331", "MATH-2473","MATH-1329"]:
        return "Calculus" 
    elif "statistics" in course_name or "stats" in course_name:
        return "Statistics"
    elif course_number in ["MATH-1315"]:
        return "College Algebra" 
    elif course_number in ["MATH-2417"]:
        return "Pre-Calculus"
    #anaysis 1 & 2, modern algebra, discrete 1 & 2, introduction to advanced mathmatics, introduction to combinitorics, Number systems
    elif course_number in ["MATH-3380", "MATH-4315", "MATH-4307", "MATH-2358","MATH-3398", "MATH-3330", "MATH-4350", "MATH-3325"]:
        return "Proof-Based Courses"
    elif course_number.startswith("MATH-"):
        return "Math (Other)"
    elif "CHEM-" in course_number or "PHYS-" in course_number or "BIO-" in course_number or "GEO-" in course_number:
        return "Science"
    else:
        return "Other"

#Catagorizes majors based on Math, Stem, non-stem
def categorize_majors(row):
    Major = row["Majors"]
    #dictionaries to hold math and stem majors
    Math_Majors = ["Applied Mathematics", "Mathematics"]
    Stem_Majors = ["Animal Science", "Aquatic Biology", "Biochemistry", "Biology", 
                   "Chemistry", "Civil Engineering","Computer Information Systems", 
                   "Computer Science", "Concrete Industry Management", "Construction Sci & Mgt", "Electrical Engineering", "Engineering Technology", 
                   "Exercise & Sports Science", "Geo Natural Resources & Envi", 
                   "Geog Resource & Enviro Stdies", "Health Sciences", "Industrial Engineering", 
                   "Manufacturing Engineering", "Mechanical Engineering", "Microbiology", 
                   "Microbiology/Molecular Genetic", "Nutrition & Foods", "Physics", "Public Health", "Radiation Therapy", 
                   "Sound Recording Technology", "Wildlife Biology"]
    if Major in Math_Majors:
        return "Math"
    elif Major in Stem_Majors:
        return "Stem" 
    else:
        return "Non-Stem"

    
df = pd.read_excel(file_input)

df["Check In Date"] = pd.to_datetime(df["Check In Date"]).dt.date

# Insert a new column called "course name" after the "course number" column   *can only insert once cause the coulmn already exists
course_num_index = df.columns.get_loc("Course Number")
# df.insert(course_num_index + 1, "Course Name", "")
df.insert(course_num_index + 1, "Course Category", "")
# add the day of the week
date_index = df.columns.get_loc("Check In Date")
df.insert(date_index + 1, "Weekday", pd.to_datetime(df["Check In Date"]).dt.day_name())
# add the hour arrived
check_in_time_index = df.columns.get_loc("Check In Time")
df.insert(check_in_time_index +1 , "Hour Arrived", "")
#add the major category
major_index = df.columns.get_loc("Majors")
df.insert(major_index + 1, "Major Category", "")

#this map is no longer needed with the new data file containing the course name *****
# Fill in course names using your mapping
# mapping_df = pd.read_csv("mapping_for_courses.csv")
# course_map = dict(zip(mapping_df["Course Number"], mapping_df["Course Name"]))
# df["Course Name"] = df["Course Number"].map(course_map)
# remove "CT" 
df["Check In Time"] = df["Check In Time"].str.replace(" CT", "")
df["Check In Time"] = pd.to_datetime(df["Check In Time"], format="%I:%M %p")
# seperate things into catagories 
df["Course Category"] = df.apply(categorize_course, axis=1)
df["Major Category"] = df.apply(categorize_majors, axis = 1)

#add hour arrived
hour_arrived = df["Check In Time"].dt.strftime("%#I %p")
df["Hour Arrived"] = hour_arrived
#fix the check in time to not have a date
df["Check In Time"] = df["Check In Time"].dt.strftime("%#I:%M %p")

# deletes the columns not needed
df = df.drop(columns = ["Categories", "Tags", "Classification", "Location", "Cumulative GPA"])
#This is used so that the last 3 rows that contain "unique studetns" is deleted
df = df[df["Course Name"].notna()]

#anonymizes student names
uniqueStudents = df["NetID"].unique()
counter = 1000
for name in uniqueStudents:
    df.loc[df["NetID"] == name, "NetID"] = counter
    counter += 1

# Save to a new Excel file
df.to_excel(file_output, sheet_name = sheetName, index=False)
# format the excel file
wb = load_workbook(file_output)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(file_output)



