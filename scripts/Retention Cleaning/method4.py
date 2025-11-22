#method 4, combines all courses a student came in for as a total occurance value
#this is used to say that a student could have worked on eny of their classes when they came in

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = "../../results/Student Course Count.xlsx"
output_file = "../../results/Retention Cleaning/method4.xlsx"
sheetName = "Retention Cleaning Method 4"

original = pd.read_excel(input_file)

course_sum = original.groupby("NetID")["Count"].sum()

final = original.copy()

final["Count"] = final["NetID"].map(course_sum)

final = final.sort_values(by="NetID", ascending=True)
final.to_excel(output_file, sheet_name=sheetName, index=False)

wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)
