#method 3, this file is used to take a students low courses and add them into their highest course
#this acounts for things like 1 or 2 time miss clicks
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = "../../results/Student Course Count.xlsx"
output_file = "../../results/Retention Cleaning/method3.xlsx"
sheetName = "Retention Cleaning Method 3"

def combine_low_courses(group):
    g = group.copy()
    #takes low counts and adds to the top course
    if (g["Count"] > 2).any():
        top = g.loc[g["Count"].idxmax()].copy()
        total_low = g.loc[g["Count"] <= 2, "Count"].sum()
        top["Count"] += total_low
        return pd.DataFrame([top])
    #if all courses are "low counts"
    else:
        math_courses = g[g["Course Category"].str.lower() == "math"]
        #there is a math course
        if not math_courses.empty:
            top = math_courses.loc[math_courses["Count"].idxmax()].copy()
            total_low = g.loc[g["Count"] <= 2, "Count"].sum()
            top["Count"] += total_low
            return pd.DataFrame([top])
        #no math course; do nothing
        else:
            return g
        
original = pd.read_excel(input_file)

final = original.groupby("NetID", group_keys=False).apply(combine_low_courses)

final = final.sort_values(by="NetID", ascending=True)
final.to_excel(output_file, sheet_name=sheetName, index=False)

wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)
