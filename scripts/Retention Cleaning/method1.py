#method one, this file will take student courses and change every arrival 
#to whichever course has the most arrivals

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = "../../results/Student Course Count.xlsx"
output_file = "../../results/Retention Cleaning/method1.xlsx"
sheetName = "Retention Cleaning Method 1" 

original = pd.read_excel(input_file)

index = original.groupby("NetID")["Count"].idxmax()
df_max = original.loc[index].copy()

total_checkins = original.groupby("NetID")["Count"].sum()

df_max["Count"] = df_max["NetID"].map(total_checkins)
# save to Excel
df_max.to_excel(output_file, sheet_name = sheetName, index=False)

# format the excel file
wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)



