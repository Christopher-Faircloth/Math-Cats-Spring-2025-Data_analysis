#method5 is a combination of method 2 first, getting rid of non math courses and adding them to their top course,
#then method4 of making all their checkins the same
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

input_file = "../../results/Student Course Count.xlsx"
output_file = "../../results/Retention Cleaning/method5.xlsx"
sheetName = "Retention Cleaning Method 5" 

original = pd.read_excel(input_file)

students_with_math = original.loc[original["Course Category"] == "Math", "NetID"].unique()

math_df = original[original["Course Category"] == "Math"]
non_math_df = original[original["Course Category"] != "Math"]

non_math_sums = (
    non_math_df[non_math_df["NetID"].isin(students_with_math)]
    .groupby("NetID", as_index=False)["Count"].sum()
    .rename(columns={"Count": "NonMath_Added"})
)

top_math = (
    math_df.sort_values(["NetID", "Count"], ascending=[True, False])
    .drop_duplicates(subset="NetID", keep="first")
)

other_math = math_df[~math_df.index.isin(top_math.index)]

top_math = top_math.merge(non_math_sums, on="NetID", how="left")
top_math["Count"] = top_math["Count"] + top_math["NonMath_Added"].fillna(0)
top_math = top_math.drop(columns=["NonMath_Added"])


final = pd.concat([
    top_math,
    other_math,
    non_math_df[~non_math_df["NetID"].isin(students_with_math)]
], ignore_index=True)

final = final.sort_values(by="NetID",ascending=True)

course_sum = final.groupby("NetID")["Count"].sum()

final["Count"] = final["NetID"].map(course_sum)

final.to_excel(output_file, sheet_name=sheetName, index=False)

wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)