#this file is used to get all the courses a single studetn came in and clean the check-in document to show all
#the courses they came in for. The reason this is necisary is to correct for when students missclick as well as
#when students come in for a class they normally work on all their classes, so although they may not select the
#same class every time they get tutored for it regardless


# as of 10/25/2025
# this file contain 812 unique ids
# this drops from 850 due to 48 students not haveing a final grade
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# this is used to change the course categories to just stem and non-stem
def change_categories(row):
    course_name = str(row["Course Number"])
    course_name = course_name.lower()
    if course_name.startswith("math-"):
        return "Math"
    else:
        return "Non-Math"

input_file = "../data/Check_In_Cleaned.xlsx"
output_file = "../results/Student Course Count.xlsx"
sheetName = "Student course count" 
original = pd.read_excel(input_file)


df = (
    original
    .groupby(['NetID', 'Course Name', 'Course Number', 'Course Category', 'Final Grade'])
    .size()
    .reset_index(name='Count')
)
df["Course Category"] = df.apply(change_categories, axis=1)
# save to Excel
df.to_excel(output_file, sheet_name = sheetName, index=False)

# format the excel file
wb = load_workbook(output_file)
ws = wb[sheetName]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheetName)
wb.save(output_file)



