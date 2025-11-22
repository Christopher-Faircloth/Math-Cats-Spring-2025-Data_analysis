import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import Counter

# input_file = "../results/Student Course Count.xlsx"
# output_file = "../results/course_retention.xlsx"
input_file = "../results/Retention Cleaning/method3.xlsx"
output_file = "../results/Retention Cleaning/method3_retention.xlsx"
sheet_name = "course retention"

def addCourseRetention(original):

    rows = []
    for course, group in original.groupby("Course Name"):
        counts = group["Count"].tolist()
        course_hist = Counter(counts)
        course_hist = dict(sorted(course_hist.items()))

        occurance_chunks = {
            "1": course_hist.get(1, 0),
            "2+": sum(v for k, v in course_hist.items() if k >= 2),
            "5+": sum(v for k, v in course_hist.items() if k >= 5),
            "10+": sum(v for k, v in course_hist.items() if k >= 10),
        }

        total_people = group["NetID"].nunique()
        retention = {
            "Course Name": course,
            "Total People": total_people,
            "1 Occurrence Count": occurance_chunks["1"],
            "1 Percent Retention": occurance_chunks["1"] / total_people if total_people else 0,
            "2+ Count": occurance_chunks["2+"],
            "2+ Percent Retention": occurance_chunks["2+"] / total_people if total_people else 0,
            "5+ Count": occurance_chunks["5+"],
            "5+ Percent Retention": occurance_chunks["5+"] / total_people if total_people else 0,
            "10+ Count": occurance_chunks["10+"],
            "10+ Percent Retention": occurance_chunks["10+"] / total_people if total_people else 0,
        }
        rows.append(retention)

    counts = original["Count"].tolist()
    overall_hist = Counter(counts)
    overall_hist = dict(sorted(overall_hist.items()))

    occurance_chunks = {
        "1": overall_hist.get(1, 0),
        "2+": sum(v for k, v in overall_hist.items() if k >= 2),
        "5+": sum(v for k, v in overall_hist.items() if k >= 5),
        "10+": sum(v for k, v in overall_hist.items() if k >= 10),
    }
    total_people = original["NetID"].nunique()
    overall_retention = {
        "Course Name": "All Courses",
        "Total People": total_people,
        "1 Occurrence Count": occurance_chunks["1"],
        "1 Percent Retention": occurance_chunks["1"] / total_people if total_people else 0,
        "2+ Count": occurance_chunks["2+"],
        "2+ Percent Retention": occurance_chunks["2+"] / total_people if total_people else 0,
        "5+ Count": occurance_chunks["5+"],
        "5+ Percent Retention": occurance_chunks["5+"] / total_people if total_people else 0,
        "10+ Count": occurance_chunks["10+"],
        "10+ Percent Retention": occurance_chunks["10+"] / total_people if total_people else 0,
    }
    rows.append(overall_retention)

    return pd.DataFrame(rows)

original = pd.read_excel(input_file)
df = addCourseRetention(original)

df.to_excel(output_file, sheet_name=sheet_name, index=False)

wb = load_workbook(output_file)
ws = wb[sheet_name]

for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_len + 2

wb.active = wb.sheetnames.index(sheet_name)
wb.save(output_file)